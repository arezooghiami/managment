from io import BytesIO

import jdatetime
import pandas as pd
from fastapi import APIRouter, Request, Depends, HTTPException
from fastapi import Form
from fastapi import UploadFile, File
from fastapi.responses import StreamingResponse
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font
from persiantools.jdatetime import JalaliDate
from sqlalchemy.orm import Session
from starlette.responses import JSONResponse, HTMLResponse, RedirectResponse

from DB.database import get_db
from models.CallEventStatus import CallEventStatus
from models.ComplaintIssue import ComplaintIssue
from models.IncomingCallEvent import IncomingCallEvent
from models.branch import Branch, Unit
from models.inCall import IncomingCall
from models.outCall import OutCall
# from models.postycode import PostyCodeStatus
from models.user import User

router_crm = APIRouter(
    tags=["Crm"],  # ← تگ دسته‌بندی در سوَگر
)
templates = Jinja2Templates(directory="templates")
from datetime import date, datetime


@router_crm.get("/crm_dashboard")
def crm_dashboard(request: Request, db: Session = Depends(get_db)):
    user_id = request.session.get("user_id")
    role = request.session.get("role")
    if not user_id:
        return RedirectResponse(url="/", status_code=302)

    user = db.query(User).filter(User.id == user_id).first()
    today = date.today()
    shamsi_today = jdatetime.date.fromgregorian(date=today).strftime('%Y/%m/%d')

    incoming_call = (
        db.query(IncomingCall)
        .filter(IncomingCall.user_id == user_id, IncomingCall.datetime == today)
        .first()
    )

    out_call = (
        db.query(OutCall)
        .filter(OutCall.user_id == user_id, OutCall.datetime == today)
        .first()
    )
    issues = db.query(ComplaintIssue).all()
    branches = db.query(Branch).all()
    units = db.query(Unit).all()

    incoming_data = {
        "posty_code": incoming_call.posty_code if incoming_call and incoming_call.posty_code else 0,
        "send_product_deadline": incoming_call.send_product_deadline if incoming_call and incoming_call.send_product_deadline else 0,
        "branch_change": incoming_call.branch_change if incoming_call and incoming_call.branch_change else 0,
        "online_change": incoming_call.online_change if incoming_call and incoming_call.online_change else 0,
        "online_return": incoming_call.online_return if incoming_call and incoming_call.online_return else 0,
        "branch_dissatisfaction": incoming_call.branch_dissatisfaction if incoming_call and incoming_call.branch_dissatisfaction else 0,
        "payment_followup": incoming_call.payment_followup if incoming_call and incoming_call.payment_followup else 0,
        "incomplete_delivery": incoming_call.incomplete_delivery if incoming_call and incoming_call.incomplete_delivery else 0,
        "b2b_sales": incoming_call.b2b_sales if incoming_call and incoming_call.b2b_sales else 0,
        "waiting_for_payment": incoming_call.waiting_for_payment if incoming_call and incoming_call.waiting_for_payment else 0,
        "product_search": incoming_call.product_search if incoming_call and incoming_call.product_search else 0,
        "after_sales_service": incoming_call.after_sales_service if incoming_call and incoming_call.after_sales_service else 0,
        "club": incoming_call.club if incoming_call and incoming_call.club else 0,
        "other": incoming_call.other if incoming_call and incoming_call.other else 0,
        "branch_info": incoming_call.branch_info if incoming_call and incoming_call.branch_info else 0,
        "product_site_info": incoming_call.product_site_info if incoming_call and incoming_call.product_site_info else 0,
        "snapp_pay": incoming_call.snapp_pay if incoming_call and incoming_call.snapp_pay else 0,
        "inner_call": incoming_call.inner_call if incoming_call and incoming_call.inner_call else 0,
        "defective_product": incoming_call.defective_product if incoming_call and incoming_call.defective_product else 0,
    }

    out_data = {
        "internet": out_call.internet if out_call and out_call.internet else 0,
        "voice_mail": out_call.voice_mail if out_call and out_call.voice_mail else 0
    }
    return templates.TemplateResponse("user/CRM.html", {
        "request": request,
        "user": user,
        "incoming_data": incoming_data,
        "out_data": out_data,
        "today": shamsi_today , # 👈 فقط تاریخ روز
        'issues':issues,
        "branches":branches,
        "units":units



    })


@router_crm.post("/update_crm_data", summary="به‌روزرسانی اطلاعات CRM")
async def update_crm_data(
        request: Request,
        db: Session = Depends(get_db)
):
    user_id = request.session.get("user_id")
    if not user_id:
        return RedirectResponse("/", status_code=302)

    try:
        data = await request.json()
    except:
        raise HTTPException(400, "Invalid JSON data")

    type_ = data.get("type")      # incoming | out
    field = data.get("field")
    change = data.get("change")
    status = data.get("status")  # only for incoming calls

    if change not in (-1, 1):
        raise HTTPException(400, "Invalid change value. Must be -1 or 1")

    today = date.today()
    now = datetime.utcnow()

    # ==================================================
    # ================= INCOMING =======================
    # ==================================================
    if type_ == "incoming":
        if not field:
            raise HTTPException(400, "Invalid field")

        # Get or create incoming call record
        incoming_call = (
            db.query(IncomingCall)
            .filter_by(user_id=user_id, datetime=today)
            .first()
        )

        if not incoming_call:
            # ایجاد رکورد جدید با مقادیر پیش‌فرض
            incoming_call = IncomingCall(
                user_id=user_id,
                datetime=today,
                start_datetime=now,
                end_datetime=now,
            )
            db.add(incoming_call)
            db.flush()

        incoming_call.end_datetime = now

        # Update the field value
        current_value = getattr(incoming_call, field, 0) or 0
        new_value = max(0, current_value + change)
        setattr(incoming_call, field, new_value)

        # Create event record for the change
        if change == 1:
            # Create new event for +1
            event = IncomingCallEvent(
                incoming_call_id=incoming_call.id,
                topic=field,
                user_id=user_id,  # فقط برای IncomingCallEvent
                created_at=now
            )
            db.add(event)
            db.flush()  # Flush to get event.id

            # Add status for the event
            # فقط فیلدهایی که در مدل CallEventStatus وجود دارند
            call_event_status = CallEventStatus(
                call_event_id=event.id,
                status=status if status else 1  # Default: حل شده
                # user_id را حذف کردیم چون در مدل وجود ندارد
            )
            db.add(call_event_status)

        else:
            # For -1 (rollback), delete the last event
            event = (
                db.query(IncomingCallEvent)
                .filter_by(
                    incoming_call_id=incoming_call.id,
                    topic=field,
                    user_id=user_id  # شرط user_id برای امنیت
                )
                .order_by(IncomingCallEvent.created_at.desc())
                .first()
            )

            if event:
                # Also delete associated status
                db.query(CallEventStatus).filter_by(call_event_id=event.id).delete()
                db.delete(event)

    # ==================================================
    # ================= OUTGOING =======================
    # ==================================================
    elif type_ == "out":
        ALLOWED_FIELDS = {"internet", "voice_mail"}
        if field not in ALLOWED_FIELDS:
            raise HTTPException(400, "Invalid out field")

        # Get or create out call record
        out_call = (
            db.query(OutCall)
            .filter_by(user_id=user_id, datetime=today)
            .first()
        )

        if not out_call:
            out_call = OutCall(
                user_id=user_id,
                datetime=today,
                internet=0,
                voice_mail=0,
            )
            db.add(out_call)
            db.flush()

        # Update the field value
        current_value = getattr(out_call, field) or 0
        new_value = max(0, current_value + change)
        setattr(out_call, field, new_value)

    else:
        raise HTTPException(400, "Invalid type. Must be 'incoming' or 'out'")

    try:
        db.commit()
        return {"success": True}
    except Exception as e:
        db.rollback()
        raise HTTPException(500, f"Database error: {str(e)}")

# این endpoint جدید برای گرفتن وضعیت‌های تماس‌ها
@router_crm.get("/get_call_statuses")
def get_call_statuses(request: Request, db: Session = Depends(get_db)):
    user_id = request.session.get("user_id")
    if not user_id:
        return RedirectResponse("/", status_code=302)

    today = date.today()

    incoming_call = (
        db.query(IncomingCall)
        .filter_by(user_id=user_id, datetime=today)
        .first()
    )

    statuses = {}
    if incoming_call:
        # گرفتن آخرین وضعیت هر تماس
        events = (
            db.query(IncomingCallEvent, CallEventStatus)
            .join(CallEventStatus, IncomingCallEvent.id == CallEventStatus.call_event_id)
            .filter(IncomingCallEvent.incoming_call_id == incoming_call.id)
            .order_by(IncomingCallEvent.created_at.desc())
            .all()
        )

        for event, status in events:
            statuses[event.topic] = status.status

    return statuses

def convert_persian_to_english_numbers(s):
    persian_nums = "۰۱۲۳۴۵۶۷۸۹"
    english_nums = "0123456789"
    for p, e in zip(persian_nums, english_nums):
        s = s.replace(p, e)
    return s


@router_crm.post("/report_crm_data")
async def report_crm_data(
        request: Request,
        jalali_date_start: str = Form(...),
        jalali_date_end: str = Form(...),
        code: str = Form(""),
        db: Session = Depends(get_db)
):
    user_id = request.session.get("user_id")
    role = request.session.get("role")
    if not user_id:
        return RedirectResponse(url="/", status_code=302)

    is_crm = request.session.get("is_crm")
    code = convert_persian_to_english_numbers(code)
    jalali_date_start = convert_persian_to_english_numbers(jalali_date_start)
    jalali_date_end = convert_persian_to_english_numbers(jalali_date_end)

    try:
        start_date = JalaliDate.strptime(jalali_date_start, "%Y/%m/%d").to_gregorian()
        end_date = JalaliDate.strptime(jalali_date_end, "%Y/%m/%d").to_gregorian()
    except:
        raise HTTPException(status_code=400, detail="فرمت تاریخ اشتباه است.")

    if role == "user" and is_crm:
        user = db.query(User).filter(User.id == user_id).first()
        if not user:
            raise HTTPException(status_code=404, detail="کاربر یافت نشد.")
        code = user.code

    users_query = db.query(User)
    if code:
        users_query = users_query.filter(User.code == code)
    users = users_query.filter(User.is_crm == True).all()

    results = []

    incoming_fields = [
        "posty_code", "send_product_deadline", "branch_change", "online_change",
        "online_return", "branch_dissatisfaction", "payment_followup", "incomplete_delivery",
        "b2b_sales", "waiting_for_payment", "product_search", "after_sales_service",
        "club", "other","branch_info","product_site_info","snapp_pay","inner_call","defective_product"
    ]
    outgoing_fields = ["internet", "voice_mail"]

    # مقدار اولیه مجموع کل همه کاربران
    total_sum_all = {field: 0 for field in incoming_fields + outgoing_fields}
    total_sum_all["total_row_sum"] = 0  # مجموع ردیفی همه کاربران

    for user in users:
        incoming_records = db.query(IncomingCall).filter(
            IncomingCall.user_id == user.id,
            IncomingCall.datetime >= start_date,
            IncomingCall.datetime <= end_date
        ).all()

        out_records = db.query(OutCall).filter(
            OutCall.user_id == user.id,
            OutCall.datetime >= start_date,
            OutCall.datetime <= end_date
        ).all()

        def sum_fields(records, fields):
            return {field: sum(getattr(r, field) or 0 for r in records) for field in fields}

        incoming_data = sum_fields(incoming_records, incoming_fields)
        out_data = sum_fields(out_records, outgoing_fields)

        # محاسبه مجموع ردیفی برای هر کاربر
        total_row_sum = sum(incoming_data.values()) + sum(out_data.values())

        # به‌روزرسانی مجموع کل همه کاربران
        for field, value in {**incoming_data, **out_data}.items():
            total_sum_all[field] += value
        total_sum_all["total_row_sum"] += total_row_sum

        results.append({
            "name": f"{user.name} {user.family}",
            "code": user.code,
            "incoming": incoming_data,
            "out": out_data,
            "total_row_sum": total_row_sum  # ستون مجموع این ردیف
        })

    # افزودن ردیف مجموع کل به انتهای خروجی
    results.append({
        "name": "مجموع کل",
        "code": "-",
        "incoming": {field: total_sum_all[field] for field in incoming_fields},
        "out": {field: total_sum_all[field] for field in outgoing_fields},
        "total_row_sum": total_sum_all["total_row_sum"]
    })

    return JSONResponse(content={"results": results})


@router_crm.get("/report", response_class=HTMLResponse)
async def get_report_page(request: Request,
                          db: Session = Depends(get_db),
                          templates: Jinja2Templates = Depends(lambda: Jinja2Templates(directory="templates"))):
    role = request.session.get("role")
    is_crm = request.session.get("is_crm")
    user_code = request.session.get("user_code")  # کد پرسنلی کاربر فعلی
    user_id = request.session.get("user_id")
    if not user_id:
        return RedirectResponse(url="/", status_code=302)

    user = db.query(User).filter(User.id == user_id).first()

    return templates.TemplateResponse("user/crm_report.html", {
        "request": request,
        "role": role,
        "is_crm": is_crm,
        "user_code": user_code,
        "user": user

    })
    # return templates.TemplateResponse("user/crm_report.html", {"request": request})


@router_crm.post("/report_crm_excel")
async def report_crm_excel(
        request: Request,
        jalali_date_start: str = Form(...),
        jalali_date_end: str = Form(...),
        code: str = Form(""),
        db: Session = Depends(get_db)
):
    user_id = request.session.get("user_id")
    role = request.session.get("role")
    is_crm = request.session.get("is_crm")
    if not user_id:
        return RedirectResponse(url="/", status_code=302)

    code = convert_persian_to_english_numbers(code)
    jalali_date_start = convert_persian_to_english_numbers(jalali_date_start)
    jalali_date_end = convert_persian_to_english_numbers(jalali_date_end)

    try:
        start_date = JalaliDate.strptime(jalali_date_start, "%Y/%m/%d").to_gregorian()
        end_date = JalaliDate.strptime(jalali_date_end, "%Y/%m/%d").to_gregorian()
    except:
        raise HTTPException(status_code=400, detail="فرمت تاریخ اشتباه است.")

    if role == "user" and is_crm:
        user = db.query(User).filter(User.id == user_id).first()
        if not user:
            raise HTTPException(status_code=404, detail="کاربر یافت نشد.")
        code = user.code

    users_query = db.query(User)
    if code:
        users_query = users_query.filter(User.code == code)
    users = users_query.filter(User.is_crm == True).all()

    incoming_fields = [
        "posty_code", "send_product_deadline", "branch_change", "online_change",
        "online_return", "branch_dissatisfaction", "payment_followup", "incomplete_delivery",
        "b2b_sales", "waiting_for_payment", "product_search", "after_sales_service",
        "club", "other","branch_info","product_site_info","snapp_pay","inner_call","defective_product"
    ]
    outgoing_fields = ["internet", "voice_mail"]

    headers = [
        "نام", "کد پرسنلی", "رهگیری", "ارسال کالا", "تعویض شعبه", "تعویض آنلاین",
        "مرجوع آنلاین", "نارضایتی شعبه", "پیگیری واریزی", "ارسال ناقص",
        "فروش سازمانی", "در انتظار پرداخت", "سرچ کالا", "پس از فروش",
        "باشگاه", "متفرقه","اطلاعات شعب","اطلاعات سایت و محصول","اسنپ ‌پی","داخلی" ,"کالای ایراد دار","پیگیری اینترنتی", "صندوق صوتی",
        "مجموع تماس‌های ورودی", "مجموع تماس‌های خروجی", "درصد تماس‌های ورودی"
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "گزارش CRM"
    ws.sheet_view.rightToLeft = True

    ws.append(headers)

    # استایل برای هدرها (پس‌زمینه رنگی + متن بولد + وسط‌چین)
    header_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    header_font = Font(bold=True, color="000000")
    header_alignment = Alignment(horizontal="center", vertical="center")
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    # استایل داده‌ها (راست‌چین + وسط عمودی)
    data_alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
        for cell in row:
            cell.alignment = data_alignment

    # تعیین عرض ستون‌ها به صورت اتوماتیک
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[col_letter].width = adjusted_width

    total_sum_all = {field: 0 for field in incoming_fields + outgoing_fields}
    total_sum_all["total_row_sum"] = 0

    for user in users:
        incoming_records = db.query(IncomingCall).filter(
            IncomingCall.user_id == user.id,
            IncomingCall.datetime >= start_date,
            IncomingCall.datetime <= end_date
        ).all()

        out_records = db.query(OutCall).filter(
            OutCall.user_id == user.id,
            OutCall.datetime >= start_date,
            OutCall.datetime <= end_date
        ).all()

        def sum_fields(records, fields):
            return {field: sum(getattr(r, field) or 0 for r in records) for field in fields}

        incoming_data = sum_fields(incoming_records, incoming_fields)
        out_data = sum_fields(out_records, outgoing_fields)

        # مجموع تماس‌های ورودی و خروجی برای هر کاربر
        incoming_sum = sum(incoming_data.values())
        outgoing_sum = sum(out_data.values())

        # مجموع کل ردیف برای هر کاربر
        total_row_sum = incoming_sum + outgoing_sum

        # بروزرسانی مجموع کل کلی
        for field, value in {**incoming_data, **out_data}.items():
            total_sum_all[field] += value
        total_sum_all["total_row_sum"] += total_row_sum

        row = [
                  f"{user.name} {user.family}",
                  user.code
              ] + [incoming_data[field] for field in incoming_fields] + \
              [out_data[field] for field in outgoing_fields] + \
              [incoming_sum, outgoing_sum, f"{round((incoming_sum / (total_row_sum or 1)) * 100, 2)}%"]

        ws.append(row)

    # محاسبه مجموع کل تماس‌های ورودی و خروجی از total_sum_all
    total_incoming_sum = sum(total_sum_all[field] for field in incoming_fields)
    total_outgoing_sum = sum(total_sum_all[field] for field in outgoing_fields)
    total_sum_all["total_row_sum"] = total_incoming_sum + total_outgoing_sum

    total_row = [
                    "مجموع کل", "-"
                ] + [total_sum_all[field] for field in incoming_fields] + \
                [total_sum_all[field] for field in outgoing_fields] + \
                [total_incoming_sum, total_outgoing_sum,
                 f"{round((total_incoming_sum / (total_sum_all['total_row_sum'] or 1)) * 100, 2)}%"]

    ws.append(total_row)

    # محاسبه درصد بر اساس توضیح شما
    total_calls_sum = total_sum_all["total_row_sum"] or 1  # ستون آخر مجموع کل
    percent_row = ["درصد", "-"]

    percent_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")  # سبز ملایم
    percent_font = Font(bold=True, color="006100")  # سبز تیره

    # ردیف درصد
    total_calls_sum = total_sum_all["total_row_sum"] or 1
    percent_row = ["درصد", "-"]

    for field in incoming_fields:
        col_sum = total_sum_all[field] or 0
        percent_value = round((col_sum / total_incoming_sum) * 100, 2) if col_sum else 0
        percent_row.append(f"{percent_value}%")  # اضافه کردن علامت %

    ws.append(percent_row)
    for cell in ws[ws.max_row]:
        cell.fill = percent_fill
        cell.font = percent_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # ساخت فایل در حافظه
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = f"crm_report_{jalali_date_start}_{jalali_date_end}.xlsx"
    response = StreamingResponse(stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response.headers["Content-Disposition"] = f"attachment; filename*=UTF-8''{filename}"
    return response


@router_crm.post("/average_report_crm")
async def average_report_crm(
        request: Request,
        jalali_date_start: str = Form(...),
        jalali_date_end: str = Form(...),
        code: str = Form(""),
        db: Session = Depends(get_db)
):
    user_id = request.session.get("user_id")
    role = request.session.get("role")
    is_crm = request.session.get("is_crm")
    if not user_id:
        return RedirectResponse(url="/", status_code=302)

    code = convert_persian_to_english_numbers(code)
    jalali_date_start = convert_persian_to_english_numbers(jalali_date_start)
    jalali_date_end = convert_persian_to_english_numbers(jalali_date_end)

    try:
        start_date = JalaliDate.strptime(jalali_date_start, "%Y/%m/%d").to_gregorian()
        end_date = JalaliDate.strptime(jalali_date_end, "%Y/%m/%d").to_gregorian()
    except:
        raise HTTPException(status_code=400, detail="فرمت تاریخ اشتباه است.")

    users_query = db.query(User)
    if code:
        users_query = users_query.filter(User.code == code)
    users = users_query.filter(User.is_crm == True).all()

    if not users:
        raise HTTPException(status_code=404, detail="کاربری یافت نشد.")

    results = []

    incoming_fields = [
        "posty_code", "send_product_deadline", "branch_change", "online_change",
        "online_return", "branch_dissatisfaction", "payment_followup", "incomplete_delivery",
        "b2b_sales", "waiting_for_payment", "product_search", "after_sales_service",
        "club", "other","branch_info","product_site_info","snapp_pay","inner_call","defective_product"
    ]
    outgoing_fields = ["internet"]

    total_sum_incoming = {field: 0 for field in incoming_fields}
    total_sum_outgoing = {field: 0 for field in outgoing_fields}
    total_sum_all = {field: 0 for field in incoming_fields + outgoing_fields}
    total_sum_all["total_row_sum"] = 0

    for user in users:
        incoming_records = db.query(IncomingCall).filter(
            IncomingCall.user_id == user.id,
            IncomingCall.datetime >= start_date,
            IncomingCall.datetime <= end_date
        ).all()

        out_records = db.query(OutCall).filter(
            OutCall.user_id == user.id,
            OutCall.datetime >= start_date,
            OutCall.datetime <= end_date
        ).all()

        def sum_fields(records, fields):
            return {field: sum(getattr(r, field) or 0 for r in records) for field in fields}

        incoming_data = sum_fields(incoming_records, incoming_fields)
        out_data = sum_fields(out_records, outgoing_fields)

        total_row_sum = sum(incoming_data.values()) + sum(out_data.values())

        for field, value in incoming_data.items():
            total_sum_incoming[field] += value
            total_sum_all[field] += value
        for field, value in out_data.items():
            total_sum_outgoing[field] += value
            total_sum_all[field] += value

        total_sum_all["total_row_sum"] += total_row_sum

        results.append({
            "name": f"{user.name} {user.family}",
            "code": user.code,
            "incoming": incoming_data,
            "out": out_data,
            "total_row_sum": total_row_sum
        })

    user_count = len(users)
    average_incoming = {field: round(value / user_count, 2) for field, value in total_sum_incoming.items()}
    average_outgoing = {field: round(value / user_count, 2) for field, value in total_sum_outgoing.items()}
    average_all = {field: round(value / user_count, 2) for field, value in total_sum_all.items()}

    return {
        "user_count": user_count,
        "total_sum_incoming": total_sum_incoming,
        "total_sum_outgoing": total_sum_outgoing,
        "total_sum_all": total_sum_all,
        "average_incoming": average_incoming,
        "average_outgoing": average_outgoing,
        "average_per_user": average_all,
        "detailed_results": results
    }


@router_crm.post("/upload_crm_excel")
async def upload_crm_excel(
        file: UploadFile = File(...),
        db: Session = Depends(get_db),
        request: Request = None
):
    # user_id = request.session.get("user_id")
    # if not user_id:
    #     return RedirectResponse(url="/", status_code=302)

    try:
        df = pd.read_excel(file.file)

        for _, row in df.iterrows():
            # استخراج نام خانوادگی از ستون اپراتور
            family_name = row['اپراتور'].strip()

            # پیدا کردن کاربر بر اساس نام خانوادگی
            user = db.query(User).filter(User.family == family_name).first()
            if not user:
                raise HTTPException(status_code=404, detail=f"User with family name {family_name} not found")

            # تبدیل تاریخ شمسی به میلادی
            shamsi_date = str(row['تاریخ']).strip()
            if '/' in shamsi_date:
                parts = shamsi_date.split('/')
                sh_year, sh_month, sh_day = map(int, parts)
                g_date = jdatetime.date(sh_year, sh_month, sh_day).togregorian()
            else:
                g_date = datetime.today().date()

            now = datetime.now()

            # IncomingCall
            incoming_fields = {
                "استعلام کد رهگیری پستی": "posty_code",
                "مهلت ارسال کالا": "send_product_deadline",
                "تعویضی مرجوعی شعب": "branch_change",
                "تعویض آنلاین": "online_change",
                "مرجوعی آنلاین": "online_return",
                "نارضایتی از شعبه": "branch_dissatisfaction",
                "پیگیری واریزی": "payment_followup",
                "ارسال ناقص": "incomplete_delivery",
                "فروش سازمانی": "b2b_sales",
                "در انتظار پرداخت": "waiting_for_payment",
                "سرچ کالا": "product_search",
                "خدمات پس از فروش": "after_sales_service",
                "باشگاه": "club",
                "متفرقه": "other",
                "اطلاعات شعب":"branch_info",
                "اطلاعات سایت و محصول":"product_site_info",
                "اسنپ‌پی":"snapp_pay",
                "داخلی":"inner_call",
                "کالای ایراد دار":"defective_product"
            }

            record = (
                db.query(IncomingCall)
                .filter(IncomingCall.user_id == user.id, IncomingCall.datetime == g_date)
                .first()
            )

            if not record:
                record = IncomingCall(
                    user_id=user.id,
                    datetime=g_date,
                    start_datetime=now,
                    end_datetime=now
                )
                db.add(record)
            else:
                # فقط زمان پایان تماس آپدیت بشه
                record.end_datetime = now

            for persian_col, db_col in incoming_fields.items():
                value = row.get(persian_col, 0)
                if value is not None:
                    setattr(record, db_col, int(value))

            # OutCall
            out_fields = {
                "پیگیری اینترنتی(تماس خروجی)": "internet",
                "پیگیری صندوق صوتی": "voice_mail"
            }

            record_out = (
                db.query(OutCall)
                .filter(OutCall.user_id == user.id, OutCall.datetime == g_date)
                .first()
            )
            if not record_out:
                record_out = OutCall(user_id=user.id, datetime=g_date)
                db.add(record_out)

            for persian_col, db_col in out_fields.items():
                value = row.get(persian_col, 0)
                if value is not None:
                    setattr(record_out, db_col, int(value))

        db.commit()
        return {"success": True, "message": "Data imported/updated successfully"}

    except Exception as e:
        return {"success": False, "error": str(e)}
