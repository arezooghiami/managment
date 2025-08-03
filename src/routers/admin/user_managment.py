from fastapi import APIRouter, Depends, Request, Form, UploadFile, File
from sqlalchemy.orm import Session
from starlette.responses import RedirectResponse, HTMLResponse
from starlette.templating import Jinja2Templates
from passlib.context import CryptContext

from DB.database import get_db
from models.office import Office
from models.user import User
from schemas.user import UserRole, Status
import openpyxl
from openpyxl.workbook import Workbook
from io import BytesIO
router = APIRouter()
templates = Jinja2Templates(directory="templates")
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

def generate_user_excel_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Users"

    # Header row
    ws.append(["نام", "نام خانوادگی", "کدپرسنلی", "پسورد",  "دفتر"])

    # Example row (optional)
    ws.append(["آرزو", "قیامی", "040571", "mypassword",  "دفتر مشهد"])

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream.read()

from fastapi.responses import StreamingResponse


@router.get("/admin/download_users_template")
def download_users_template():
    content = generate_user_excel_template()
    return StreamingResponse(
        BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=users_template.xlsx"}
    )
@router.get("/admin/users", response_class=HTMLResponse)
async def users_management(request: Request, db: Session = Depends(get_db)):
    user_id = request.session.get("user_id")
    role = request.session.get("role")

    if not user_id or role != "admin":
        return RedirectResponse(url="/", status_code=302)

    users = db.query(User).all()
    offices = db.query(Office).all()

    return templates.TemplateResponse("admin/users_management.html", {
        "request": request,
        "users": users,
        "offices": offices,
        "UserRole": UserRole,
        "Status": Status
    })

@router.post("/admin/add_user", response_class=HTMLResponse)
async def add_user(
        request: Request,
        name: str = Form(...),
        family: str = Form(...),
        code: str = Form(...),
        password: str = Form(...),
        role: UserRole = Form(...),
        office_id: int = Form(...),
        db: Session = Depends(get_db)
):
    # Check if user code already exists
    existing_user = db.query(User).filter(User.code == code).first()
    if existing_user:
        request.session["flash"] = "کد پرسنلی تکراری است"
        return RedirectResponse(url="/admin/users", status_code=303)

    # Check if office exists
    office = db.query(Office).get(office_id)
    if not office:
        request.session["flash"] = "دفتر انتخاب شده وجود ندارد"
        return RedirectResponse(url="/admin/users", status_code=303)

    # Create new user with hashed password
    hashed_password = pwd_context.hash(password)
    new_user = User(
        name=name,
        family=family,
        code=code,
        password=hashed_password,
        role=role,
        status=Status.ACTIVE,
        office_id=office_id
    )

    db.add(new_user)
    db.commit()

    request.session["flash"] = "کاربر با موفقیت افزوده شد"
    return RedirectResponse(url="/admin/users", status_code=303)

@router.post("/admin/upload_users_excel", response_class=HTMLResponse)
async def upload_users_excel(
        request: Request,
        file: UploadFile = File(...),
        db: Session = Depends(get_db)
):
    if not file.filename.endswith((".xlsx", ".xls")):
        request.session["flash"] = "فرمت فایل باید اکسل باشد"
        return RedirectResponse(url="/admin/users", status_code=303)

    try:
        content = await file.read()
        workbook = openpyxl.load_workbook(filename=BytesIO(content))
        sheet = workbook.active

        added = 0
        skipped = 0

        for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                name, family, code, password, office_name = row

                if not all([name, family, code, password, office_name]):
                    skipped += 1
                    continue

                # بررسی تکراری نبودن کد
                if db.query(User).filter(User.code == str(code)).first():
                    skipped += 1
                    continue

                # پیدا کردن دفتر
                office = db.query(Office).filter(Office.name == str(office_name)).first()
                if not office:
                    skipped += 1
                    continue

                hashed_password = pwd_context.hash(password)

                new_user = User(
                    name=name,
                    family=family,
                    code=str(code),
                    password=hashed_password,
                    role=UserRole.USER,  # یا اگر همیشه یوزر هست
                    status=Status.ACTIVE,
                    office_id=office.id
                )

                db.add(new_user)
                db.flush()  # بلافاصله اجرا کن برای گرفتن خطا

                added += 1

            except Exception as e:
                db.rollback()  # اگر در این کاربر خطا بود، عقب‌گرد
                print(f"خطا در ردیف {idx}: {e}")
                skipped += 1

        db.commit()
        request.session["flash"] = f"{added} کاربر افزوده شد، {skipped} مورد رد شد"

    except Exception as e:
        db.rollback()
        request.session["flash"] = f"خطا در پردازش فایل: {str(e)}"

    return RedirectResponse(url="/admin/users", status_code=303)

@router.post("/admin/edit_user/{user_id}", response_class=HTMLResponse)
async def edit_user(
        request: Request,
        user_id: int,
        name: str = Form(...),
        family: str = Form(...),
        code: str = Form(...),
        status: Status = Form(...),
        password: str = Form(None),  # Password is optional
        role: UserRole = Form(...),
        office_id: int = Form(...),
        db: Session = Depends(get_db)
):
    # Check admin access
    if request.session.get("role") != "admin":
        return RedirectResponse(url="/", status_code=302)

    # Check if user exists
    user = db.query(User).get(user_id)
    if not user:
        request.session["flash"] = "کاربر یافت نشد"
        return RedirectResponse(url="/admin/users", status_code=303)

    # Check if office exists
    office = db.query(Office).get(office_id)
    if not office:
        request.session["flash"] = "دفتر انتخاب شده وجود ندارد"
        return RedirectResponse(url="/admin/users", status_code=303)

    # Check if code is unique (excluding current user)
    existing_user = db.query(User).filter(User.code == code, User.id != user_id).first()
    if existing_user:
        request.session["flash"] = "کد پرسنلی تکراری است"
        return RedirectResponse(url="/admin/users", status_code=303)

    # Update user fields
    user.name = name
    user.family = family
    user.code = code
    user.status = status
    user.role = role
    user.office_id = office_id
    if password:  # Update password only if provided
        user.password = pwd_context.hash(password)

    db.commit()
    request.session["flash"] = "کاربر با موفقیت ویرایش شد"
    return RedirectResponse(url="/admin/users", status_code=303)

@router.post("/admin/delete_user/{user_id}", response_class=HTMLResponse)
async def delete_user(request: Request, user_id: int, db: Session = Depends(get_db)):
    # Check admin access
    if request.session.get("role") != "admin":
        return RedirectResponse(url="/", status_code=302)

    user = db.query(User).get(user_id)
    if not user:
        request.session["flash"] = "کاربر یافت نشد"
        return RedirectResponse(url="/admin/users", status_code=303)

    db.delete(user)
    db.commit()
    request.session["flash"] = "کاربر با موفقیت حذف شد"
    return RedirectResponse(url="/admin/users", status_code=303)